import json
from typing import Dict, Any
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import pandas as pd
import xlwings as xw    

from config import GEMINI_API_KEY

from openai import OpenAI
import google.generativeai as genai


from pathlib import Path
BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "data" / "valuations"



def load_valuation_excel(
    excel_path: str,
    sheet_name: str = "stock_val",
) -> Dict[str, Any]:

    wb = load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name]

    tmp = {}

    # Only fundamentals are read from Excel (max 35 rows)
    for row in ws.iter_rows(
        min_row=1,
        max_row=35,
        min_col=1,
        max_col=3,
        values_only=True,
    ):
        label, mid, good = row
        if label is None:
            continue
        label = str(label).strip()
        if good is None or good == "":
            good = mid
        tmp[label] = {"mid": mid, "good": good}

    def mid(label: str):
        return tmp.get(label, {}).get("mid")

    fundamentals = {
        "ticker": mid("Ticker"),
        "share_price": mid("Share Price"),
        "shares_outstanding": mid("Shares Outstanding"),
        "market_cap": mid("Market Cap (auto)"),
        "revenue_qtr": mid("Revenue (Qtr)"),
        "cogs": mid("COGS"),
        "gross_profit": mid("Gross Profit"),
        "gross_margin": mid("Gross Margin"),
        "opex": mid("OPEX"),
        "operating_profit": mid("Operating Profit"),
        "operating_margin": mid("Operating Margin"),
        "ebitda_ps": mid("EBITDA PS"),
        "cash": mid("Cash"),
        "debt": mid("Debt"),
        "net_cash": mid("Net Cash (auto)"),
    }

    scenarios = {
        "expected_rev_cagr_5y": {"label": "Expected Revenue CAGR (5y)", "mid": None, "good": None},
        "expected_op_margin": {"label": "E Operated Margin", "mid": None, "good": None},
        "expected_dilution": {"label": "E Dilution (5yr)", "mid": None, "good": None},

        "lt_net_debt": {"label": "LT Net Debt", "mid": None, "good": None},
        "interest_rate_debt": {"label": "Interest Rate on Debt", "mid": None, "good": None},
        "tax_rate": {"label": "Tax Rate", "mid": None, "good": None},

        "lt_earning_multiple": {"label": "LT Earning Multiple", "mid": None, "good": None},
    }

    return {
        "fundamentals": fundamentals,
        "scenarios": scenarios,
    }


def generate_llm_investment_summary(
    context: Dict[str, Any],
    provider: str = "gemini",
    model: str = "gemini-2.5-pro",
) -> str:

    fundamentals = context["fundamentals"]
    scenarios = context["scenarios"]

    fundamentals_json = json.dumps(fundamentals, indent=2, default=str)
    scenarios_json = json.dumps(scenarios, indent=2, default=str)

    ticker = fundamentals.get("ticker") or ""


    user_prompt = f"""

Use Yahoo Finance, Google Finance, or other web sources to get a brief overview of the company with ticker {ticker}.
Make sure that ticker is when you describe the company.

Ticker: {ticker}

FUNDAMENTALS: (values are in thousands USD, except per share and percentages)
{fundamentals_json}

SCENARIOS:
{scenarios_json}

Write a Markdown report with:

### 1. Company snapshot 
 (use your own reasoning or web sources, do not just repeat the fundamentals):
### 2. Pros 
 (use your own reasoning or web sources, do not just repeat the fundamentals):
### 3. Cons 
 (use your own reasoning or web sources, do not just repeat the fundamentals):
### 4. Scenario Suggestions 
 (Suggest two values for case mid and good case for each input, very briefly -one sentence- explain why.)
### Include the suggested two values in your explanation also the same in json file at the end. 
scenario inputs are these:
- Expected Revenue CAGR (5y) (expected_rev_cagr_5y)
- Expected Operating Margin (expected_op_margin)
- Expected Dilution (5y) (expected_dilution)
- Longterm Debt (lt_net_debt)
- Interest Rate on Debt (interest_rate_debt)
- Tax rate (tax_rate)
- Long Term Earning Multiple (lt_earning_multiple)


Keep the whole report concise less than 200 words.

AFTER you finish the report, on a new line write exactly:
SCENARIO_JSON_START

On the next line output ONLY a valid JSON object with numeric mid/good values for each scenario key,
with this exact structure (percentages as decimals, e.g. 0.35 for 35%):

{{
  "expected_rev_cagr_5y": {{ "mid": 0.35, "good": 0.50 }},
  "expected_op_margin": {{ "mid": 0.25, "good": 0.35 }},
  "expected_dilution": {{ "mid": 0.05, "good": 0.10 }},
  "lt_net_debt": {{ "mid": 0, "good": 0 }},
  "interest_rate_debt": {{ "mid": 0.05, "good": 0.06 }},
  "tax_rate": {{ "mid": 0.20, "good": 0.20 }},
  "lt_earning_multiple": {{ "mid": 20, "good": 25 }}
}}

Use reasonable values instead of the example above.
Do NOT wrap this JSON in markdown code fences and do NOT add any text after the JSON.


"""
    # print(user_prompt)
    
    if provider.lower() == "gemini":

        genai.configure(api_key=GEMINI_API_KEY)

        # Default Gemini model
        if model is None:
            model = "gemini-2.5-flash"

        gm = genai.GenerativeModel(model)

        resp = gm.generate_content(user_prompt)
        return resp.text

    # OPENAI MODE
    if provider.lower() == "openai":
        
        client = OpenAI(api_key=OPENAI_API_KEY)

        # Default OpenAI model
        if model is None:
            model = "gpt-4.1-mini"

        resp = client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": "You are an equity analyst."},
                {"role": "user", "content": user_prompt},
            ],
            #temperature=0.3,
        )


    return resp.choices[0].message.content




def write_llm_result_to_excel(
    excel_path: str,
    ticker: str,
    llm_text: str,
    output_path: str = None,
    sheet_name: str = "stock_val",
):

    marker = "SCENARIO_JSON_START"
    if marker not in llm_text:
        raise ValueError("SCENARIO_JSON_START marker not found in LLM output.")

    # Marker'dan sonrasını al
    text_part, json_part = llm_text.split(marker, 1)
    json_str = json_part.strip()

    scenario_json = json.loads(json_str)

    # Write text_part to Excel
    wb = load_workbook(excel_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {excel_path}")

    ws = wb[sheet_name]

    ws.merge_cells("F3:K40")
    cell = ws["F3"]
    cell.value = text_part
    cell.alignment = Alignment(wrap_text=True, vertical="top")

    scenario_labels = {
    "expected_rev_cagr_5y": "Expected Revenue CAGR (5y)",
    "expected_op_margin": "E Operated Margin",
    "expected_dilution": "E Dilution (5yr)",
    "lt_net_debt": "LT Net Debt",
    "interest_rate_debt": "Interest Rate on Debt",
    "tax_rate": "Tax Rate",
    "lt_earning_multiple": "LT Earning Multiple",
}
    
    # A sütununda etiketleri bul, B=Mid, C=Good
    max_row = ws.max_row or 100
    for key, label in scenario_labels.items():
        if key not in scenario_json:
            continue

        mid_val = scenario_json[key].get("mid")
        good_val = scenario_json[key].get("good", mid_val)

        # İlgili label satırını bul
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=3):
            label_cell = row[0]  # A sütunu
            if label_cell.value is None:
                continue
            if str(label_cell.value).strip() == label:
                mid_cell = row[1]   # B sütunu
                good_cell = row[2]  # C sütunu
                mid_cell.value = mid_val
                good_cell.value = good_val
                break

    # Determine output filename
    if output_path is None:
        import os
        folder = os.path.dirname(excel_path)
        output_path = os.path.join(folder, f"ai-summaries", f"{ticker}_ai.xlsx")

    wb.save(output_path)

    # Remove original excel
    try:
        os.remove(excel_path)
    except OSError:
        pass

    # Read with pandas to read formula values
    try:
        app = xw.App(visible=False)
        book = app.books.open(os.path.abspath(output_path))
        book.save()
        app.quit()
    
    except Exception as e:
        print(f"Warning: xlwings failed: {e}") 

    df_pred = pd.read_excel(output_path, header=None)
    

    predictions = {
        "ticker": ticker,
        "current_price": f"{df_pred.iloc[3, 1]:.2f}",      # Row 4, Column B (index 3, 1)
        "lower_prediction": f"{df_pred.iloc[37, 1]:.2f}",  # Row 38, Column B (index 37, 1)
        "upper_prediction": f"{df_pred.iloc[37, 2]:.2f}"   # Row 38, Column C (index 37, 2)
    }

    return text_part, predictions, output_path





if __name__ == "__main__":
    excel_file = DATA_DIR / "AAPL.xlsx"


