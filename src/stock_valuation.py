from openpyxl import load_workbook
from src.fin_data_yf import YFinanceDataFetcher
import os
import pandas as pd
import numpy as np
import xlwings as xw


def value_stock(ticker: str, save_file:bool = True, 
                   template_path: str ="./data/format.xlsx", 
                   output_dir: str = "./data/valuations",
                   api_source: str ="YF"):
    

    if api_source == "YF":
        fetcher = YFinanceDataFetcher(ticker)
    else:
        raise ValueError("api_source must be 'YF")
    
    data = fetcher.fetch_all_data()
    
    wb = load_workbook(template_path)
    ws = wb.active
    ws['B3'] = ticker
    # Map data to Excel rows (Column C = values)
    row_mapping = {
        4: 'Share Price',
        5: 'Shares Outstanding',
        7: 'Revenue (Qtr)',
        8: 'COGS',
        12: 'OPEX',
        13: 'Operating Profit',
        17: 'Cash',
        18: 'Debt',
    }
    
    for row, key in row_mapping.items():
        if key in data:
            ws[f'B{row}'] = data[key]
    
        
    # Save as {ticker}.xlsx in data/valuation
    if save_file:
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, f"{ticker}.xlsx")
        wb.save(output_path)
        print(f"Saved to {output_path}")

    else:
        # Save to temp location to calculate formulas
        output_path = "./data/valuations/temp_calc.xlsx"
        os.makedirs("./data/valuations", exist_ok=True)
        wb.save(output_path)
      
    try:
        app = xw.App(visible=False)
        book = app.books.open(os.path.abspath(output_path))
        book.save()
        app.quit()
    
    except Exception as e:
        print(f"Warning: xlwings failed: {e}")    
    
    # Read with pandas to read formula values
    df_pred = pd.read_excel(output_path, header=None)
    

    predictions = {
        "ticker": ticker,
        "current_price": f"{df_pred.iloc[3, 1]:.2f}",      # Row 4, Column B (index 3, 1)
        "lower_prediction": f"{df_pred.iloc[37, 1]:.2f}",  # Row 38, Column B (index 37, 1)
        "upper_prediction": f"{df_pred.iloc[37, 2]:.2f}"   # Row 38, Column C (index 37, 2)
    }


    return predictions


if __name__ == "__main__":
    value_stock("GOOGL", "data/format.xlsx")